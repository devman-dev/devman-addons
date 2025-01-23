from odoo import fields, models, api
from odoo.exceptions import UserError

import base64
import io
from openpyxl import load_workbook

class BankMoveImported(models.Model):
    _name = 'bank.move.imported'
    _description = 'Bank Move Imported'
    
    customer_id = fields.Many2one('res.partner', string='Cliente', required=True)
    date = fields.Date(string='Fecha', required=True)
    bank_id = fields.Many2one('account.bank.pagoflex', string='Banco', required=True)
    file = fields.Binary(string='Archivo', required=True)
    comment = fields.Text(string='Comentario')
    amount = fields.Char(string='Col Monto', required=True)
    origin_account = fields.Char(string='Col Cuenta Origen')
    origin_cuit = fields.Char(string='Col Cuit Origen')
    origin_cvu = fields.Char(string='Col Cvu Origen')
    bank_commission_entry = fields.Float(string='Comisión del banco por ingreso')
    bank_commission_egress = fields.Float(string='Comisión del banco por egreso')
    service_id = fields.Many2one('collection.services.commission', string='Servicio', required=True)
    commission = fields.Float(string='Comisión (%)', digits=(16, 3))
    operation_id = fields.Many2one('product.template', relation='operation', string='Operación')
    destination_account_id = fields.Many2one('collection.services.commission', string='Cuenta Destino')
    withdrawal_operations = fields.Many2many('product.template', domain=[('collection_type', '=', 'operation')])
    collection_trans_type = fields.Selection(
        [('movimiento_recaudacion', 'Acreditación'), ('retiro', 'Mov. Retiro'), ('movimiento_interno', 'Mov. Interno')],
        default='movimiento_recaudacion',
        string='Tipo de Transacción',
    )
    cuit_destination_account = fields.Char('CUIT Destino')
    cbu_destination_account = fields.Char(string='CBU Destino', tracking=True, default=False)
    cvu_destination_account = fields.Char(string='CVU Destino', tracking=True, default=False)
    alias_destination_account = fields.Char(string='Alias Destino')
    
    @api.onchange('destination_account_id')
    def get_destination_account_data(self):
        conciliation_wiz = self.env.context.get('conciliation_wiz', False)
        if conciliation_wiz:
            return
        if self.destination_account_id:
            self.sudo().write(
                {
                    'cuit_destination_account': self.destination_account_id.cuit,
                    'cbu_destination_account': self.destination_account_id.cbu,
                    'cvu_destination_account': self.destination_account_id.cvu,
                    'alias_destination_account': self.destination_account_id.alias,
                }
            )
        else:
            self.sudo().write(
                {
                    'cuit_destination_account': '',
                    'cbu_destination_account': '',
                    'cvu_destination_account': '',
                    'alias_destination_account': '',
                }
            )

    @api.onchange('collection_trans_type', 'service')
    def set_default_operation(self):
        for rec in self:
            if rec.collection_trans_type == 'movimiento_recaudacion':
                accreditation = rec.operation_id.search([('check_accreditation', '=', True), ('collection_type', '=', 'operation')])
                if accreditation:
                    rec.withdrawal_operations = accreditation.ids
                else:
                    rec.withdrawal_operations = accreditation

            elif rec.collection_trans_type == 'retiro':
                extraction = rec.operation_id.search([('check_withdrawal', '=', True), ('collection_type', '=', 'operation')])
                services = rec.service_id.search([('services', '=', rec.service_id.services.id), ('name_account', '!=', False)])
                if extraction:
                    rec.withdrawal_operations = extraction.ids
                else:
                    rec.withdrawal_operations = extraction
                if services:
                    rec.origin_account_table = services.ids


            elif rec.collection_trans_type == 'movimiento_interno':
                internal = rec.operation_id.search([('check_internal', '=', True), ('collection_type', '=', 'operation')])
                if internal:
                    rec.withdrawal_operations = internal.ids
                else:
                    rec.withdrawal_operations = internal
    
    def execute_bank_file(self):
        file = base64.b64decode(self.file)
        excel_file = io.BytesIO(file)
        wb = load_workbook(filename=excel_file, data_only=True)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        raise UserError(data) 